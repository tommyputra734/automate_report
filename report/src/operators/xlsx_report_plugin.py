#%%
import pandas as pd 
from openpyxl import load_workbook 
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import string
import logging


class ExcelReportPlugin():
    def __init__(self, input_data, output_data):
        self.input_data = input_data
        self.output_data = output_data
    
    def main(self):
        df = self.read_input_file()
        df['Date'] = pd.to_datetime(df['Date']).dt.strftime("%Y-%m-%d")
        df_transform = self.transform(df)
        self.create_output_file(df_transform)

        wb = load_workbook(self.output_data)
        wb.active = wb['Report']

        min_column = wb.active.min_column
        max_column = wb.active.max_column
        min_row = wb.active.min_row
        max_row = wb.active.max_row
        
        self.column_dimension(wb.active)
        self.barchart(wb.active, min_column, max_column, min_row, max_row)
        self.add_total(max_column, max_row, min_row, wb.active)
        self.save_file(wb)
        
    def read_input_file(self):
        df = pd.read_excel(self.input_data)
        df['Date'] = pd.to_datetime(df['Date']).dt.strftime("%Y-%m-%d")
        logging.info(df.head())
        return df
    
    def transform(self, df:pd.DataFrame) -> pd.DataFrame:
        df_transform = df.pivot_table(index=['Gender', 'Date'],
                                      columns='Product line', 
                                      values='Total', 
                                      aggfunc='sum').round()
        return df_transform
    
    def create_output_file(self, df_transfom):
        df_transfom.to_excel(self.output_data,
                             sheet_name='Report',
                             startrow=4)
    
    def column_dimension(self, wb):
        dim_holder = DimensionHolder(worksheet=wb)

        for col in range(wb.min_column, wb.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(wb, min=col, max=col, width=20)

        wb.column_dimensions = dim_holder
        
    def barchart(self, wb, min_column, max_column, min_row, max_row):
        barchart = BarChart()

        data = Reference(wb,
                 min_col=min_column+2,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row)

        categories = Reference(wb,
                        min_col=min_column,
                        max_col=min_column+1,
                        min_row=min_row+1,
                        max_row=max_row
                        )

        barchart.add_data(data, titles_from_data=True)
        barchart.set_categories(categories)


        wb.add_chart(barchart, 'J5')
        barchart.title = 'Sales Berdasarkan Produk Perhari'
        barchart.style = 2
        barchart.width = 58
        barchart.height = 15
        
    def add_total(self, max_column, max_row, min_row, wb):
        alphabet = list(string.ascii_uppercase)
        alphabet_excel = alphabet[:max_column]
        for i in alphabet_excel:
            if i != 'A' and i != 'B':
             wb[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
             wb[f'{i}{max_row+1}'].style = 'Currency'

        wb[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

        wb['A1'] = 'Sales Report'
        wb['A2'] = '2019'
        wb['A1'].font = Font('Arial', bold=True, size=20)
        wb['A2'].font = Font('Arial', bold=True, size=10)
            
    def save_file(self, wb):
        wb.save(self.output_data)
        
#%%
import os

base_path = os.sep.join(os.getcwd().split(os.sep)[:-3])
print(f'base path: {base_path}')

input_data = base_path + '/input_data/supermarket_sales.xlsx'
output_data = base_path + '/output_data/daily_report_2.xlsx'

excelReport = ExcelReportPlugin(input_data, output_data)
excelReport.main()
# %%
