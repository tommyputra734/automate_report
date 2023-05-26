import pandas as pd 
from openpyxl import load_workbook 
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
import json

input_data = 'input_data/supermarket_sales.xlsx'
output_data = 'output_data/daily_report.xlsx'

#PART 1 - Load Dataset

df = pd.read_excel(input_data)
df['Date'] = pd.to_datetime(df['Date']).dt.strftime("%Y-%m-%d")

df = df.pivot_table(index=['Gender', 'Date'], columns='Product line', values='Total', aggfunc='sum').round()

print('save dataframe to excel...')

df.to_excel(output_data, sheet_name='Report', startrow=4)

#PART 2 - Grafik

wb = load_workbook(output_data)
wb.active = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

dim_holder = DimensionHolder(worksheet=wb.active)

for col in range(min_column, max_column + 1):
    dim_holder[get_column_letter(col)] = ColumnDimension(wb.active, min=col, max=col, width=20)

wb.active.column_dimensions = dim_holder

#barchart
barchart = BarChart()

data = Reference(wb.active,
                 min_col=min_column+2,
                 max_col=max_column,
                 min_row=min_row,
                 max_row=max_row)

categories = Reference(wb.active,
                        min_col=min_column,
                        max_col=min_column+1,
                        min_row=min_row+1,
                        max_row=max_row
                        )

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)


wb.active.add_chart(barchart, 'J5')
barchart.title = 'Sales Berdasarkan Produk Perhari'
barchart.style = 2
barchart.width = 58
barchart.height = 15
wb.save(output_data)

#%%

#total penjualan
import string
alphabet = list(string.ascii_uppercase)
alphabet_excel = alphabet[:max_column]

for i in alphabet_excel:
    if i != 'A' and i != 'B':
        wb.active[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
        wb.active[f'{i}{max_row+1}'].style = 'Currency'

wb.active[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

wb.active['A1'] = 'Sales Report'
wb.active['A2'] = '2019'
wb.active['A1'].font = Font('Arial', bold=True, size=20)
wb.active['A2'].font = Font('Arial', bold=True, size=20)

wb.save(output_data)
# %%
